import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os
import seaborn as sns
import glob # for directory parsing
from openpyxl import load_workbook  # attempt 2 at fast loading excel worksheets
from datetime import date
import math

#TODO
# 2. Allow the y-axis to contain values other than Age
# 3. Allow scripting of all values entered

# CUSTOMISED PARAMETERS (edit as needed)
# polynomial_order = 1  # order of the polynomial fit
age_delta = 5  # trends for ages for every 5 years (15-19, 20-24, etc)
age_delta_half = 2.5  # half of the age to calculate the mid point
percentiles = [0, 10, 20, 30, 40, 50, 60, 70, 80, 90, 100]

################# FUNCTIONS #######################################################

def Get_RSquared(x_list, y_list):
    n = len(x_list)
    x_bar = sum(x_list)/n
    y_bar = sum(y_list)/n
    x_std = math.sqrt(sum([(xi-x_bar)**2 for xi in x_list])/(n-1))
    y_std = math.sqrt(sum([(yi-y_bar)**2 for yi in y_list])/(n-1))
    zx = [(xi-x_bar)/x_std for xi in x_list]
    zy = [(yi-y_bar)/y_std for yi in y_list]
    r = sum(zxi*zyi for zxi, zyi in zip(zx, zy))/(n-1)

    return r**2

def PlotTrimmings(my_tab_sheet):
    # Plot the complete figure and save
    plt.legend(loc='upper right', borderaxespad=0.2, fontsize=8)
    plt.xlabel("Patient Age")
    plt.ylabel(attr_formatted)
    plt.xlim(left=min_age, right=max_age)
    if (my_tab_sheet != ""):
        plt.title(f"{my_tab_sheet}: {sex.title()} {attr_formatted} - Poly:{polynomial_order} @{date.today()}")
    else:
        plt.title(f"{sex.title()} {attr_formatted} Poly :{polynomial_order} @{date.today()}")

    plt.grid(which='major', linestyle='-', color='black')
    plt.grid(which='minor', linestyle='--')
    plt.minorticks_on()
    plt.tight_layout()
    return plt

def PlotEquation(plt, rsqr):
    box_color = "white"
    x1 = 0.007
    y1 = 0.988

    plt.text(
        x1,
        y1,
        f'r^2 = {rsqr}',
        fontsize=9,
        horizontalalignment="left",
        verticalalignment="top",
        transform=plt.gca().transAxes,
        bbox=dict(facecolor=box_color, edgecolor="lightgrey", boxstyle="round,pad=0.4", alpha=0.8),
    )

def PlotFakePatient(plt, current_patient):
    # Add a comparing the current patient
    # Calculate the patient's age bin
    current_patient["age_bin"] = pd.cut(
        [current_patient["patient_age"]],
        bins=age_bins,
        labels=age_bin_labels,
        right=False
    )[0]

    # Find the age_bin data for the current patient
    current_patient_age_bin_data = age_bin_percentiles_df[
        age_bin_percentiles_df["age_bin"] == current_patient["age_bin"]]

    # Calculate outcome measurement using percentiles and then box colour
    outcome = None
    box_color = "white"
    for idx, p in enumerate(percentiles[:-1]):
        if current_patient[attr] <= age_bin_percentiles[idx + 1](current_patient["patient_age"]):
            outcome = f"{percentiles[idx]}-{percentiles[idx + 1]}th Percentile"
            box_color = colors[idx]
            break
    if outcome is None:
        outcome = "100th Percentile"
        box_color = colors[-1]

    # Create the patient information box and plot
    patient_info = (
        f"Current Patient:\n"
        f"Sex: {current_patient['patient_sex']}\n"
        f"Age: {current_patient['patient_age']}\n"
        f"{attr_formatted}: {current_patient[attr]}\n"
        f"Outcome: {outcome}"
    )
    x1 = 0.4
    y1 = 0.98

    plt.text(
        x1,
        y1,
        patient_info,
        fontsize=9,
        horizontalalignment="left",
        verticalalignment="top",
        transform=plt.gca().transAxes,
        bbox=dict(facecolor=box_color, edgecolor="black", boxstyle="round,pad=0.2", alpha=0.3),
    )

    # Add an arrow from the patient information box to the current patient dot
    x2 = x1 + 0.05
    y2 = y1 - 0.16
    arrow_start = (x2, y2)
    arrow_end = (current_patient["patient_age"], current_patient[attr])
    plt.annotate("",
                 xy=arrow_end,
                 xycoords='data',
                 xytext=arrow_start,
                 textcoords='axes fraction',
                 arrowprops=dict(facecolor='black', arrowstyle='->', lw=1),
                 )

    # show the red bullseye
    plt.scatter(current_patient["patient_age"], current_patient[attr], color='red', edgecolors='black', linewidths=0.75)

    return plt

def FormatEquation(age_bin_percentiles, polynomial_order):
    if (polynomial_order == 1):    ## linear
        res = f"{round(age_bin_percentiles[1], 4)}x + {round(age_bin_percentiles[0], 3)}"
    elif (polynomial_order == 2):
        res = f"{round(age_bin_percentiles[2], 4)}x^2  + {round(age_bin_percentiles[1], 4)}x + {round(age_bin_percentiles[0], 3)}"
    elif (polynomial_order == 3):
        res = f"{round(age_bin_percentiles[3], 4)}x^3 + {round(age_bin_percentiles[2], 4)}x^2 + {round(age_bin_percentiles[1], 4)}x + {round(age_bin_percentiles[0], 3)}"
    else:
        res = ""

    return res

def PlotTrendLine(plt):
    # Plot median trendline
    plt.plot(x_plot,  # age_bin_percentiles_df["mid_age"],
             age_bin_percentiles[5](x_plot),  # ge_bin_percentiles_df[50],
             color="black",
             linestyle="--",
             linewidth=2.5,
             #label="Median Trendline")
             label = FormatEquation(age_bin_percentiles[5], polynomial_order)
)

    #calculate R^2
    rsqr = Get_RSquared(x_plot, age_bin_percentiles[5](x_plot))

    return rsqr, plt

def PlotScatterPlot(plt, filtered_data):
    # Plot the scatter plot dots
    sns.scatterplot(data=filtered_data,
                    x="patient_age",
                    y=attr, alpha=1.0,
                    color='black',
                    marker='.',
                    size=0.5,
                    edgecolors='none',
                    linewidth=0,
                    legend=False)
    #plt.scatter(current_patient["patient_age"], current_patient[attr], color='red', edgecolors='black', linewidths=0.75)
    return plt

def PlotPercentileRanges(polynomial_order):
    # Plot the percentile ranges with corresponding colors
    plt.figure(figsize=(10, 6))  # dimensions of the printed plot
    for i, p in enumerate(percentiles[:-1]):
        #new = f"{round(age_bin_percentiles[i][1], 3)}x + {round(age_bin_percentiles[i][0], 3)}"
        plt.fill_between(
            x_plot,
            age_bin_percentiles[i](x_plot),
            age_bin_percentiles[i + 1](x_plot),
            color=colors[i],
            linewidth=0.4,
            linestyle="-",
            edgecolor='grey',
            # label=f"{p}-{percentiles[i + 1]}th Percentile" if i > 0 else f"{p}-{percentiles[i + 1]}th Percentile",
            #label=f"{p}th {age_bin_percentiles[i]}",
            label=FormatEquation(age_bin_percentiles[i], polynomial_order),
            #label=xxx,
        )

    return plt

def CalculatePercentiles(filtered_data):
    # Define the percentiles and age bins
    # Percentiles are the deviation from the median (0, 10, 20, 30, etc)
    # Bins are how the y-axis is separated into groups (eg. if age, age bins of 5 are 15-19, 20-24, 25, 29, etc)
    # Bins are used to calculated percentiles in those bins (eg. 10% deviation from 15-19 year age bracket)
    # equation = ""
    age_bins = range(min_age, max_age + 1, age_delta)
    age_bin_labels = [f"{age}-{age + age_delta - 1}" for age in age_bins[:-1]]
    filtered_data["age_bin"] = pd.cut(filtered_data["patient_age"], bins=age_bins, labels=age_bin_labels, right=False)

    # Calculate binned average the percentile level points to create a smoother percentile level
    age_bin_percentiles = []
    for age_bin in age_bin_labels:
        age_bin_data = filtered_data[filtered_data["age_bin"] == age_bin][attr]
        age_bin_percentiles.append([age_bin] + [np.percentile(age_bin_data, p) for p in percentiles])
    age_bin_percentiles_df = pd.DataFrame(age_bin_percentiles, columns=["age_bin"] + percentiles)
    age_bin_percentiles_df["mid_age"] = [int(x.split('-')[0]) + age_delta_half for x in
                                         age_bin_percentiles_df["age_bin"]]

    # Fit a polynomial to the binned and averaged percentile level points
    age_bin_percentiles = []
    for p in percentiles:
        x = age_bin_percentiles_df["mid_age"]
        y = [np.percentile(filtered_data[filtered_data["age_bin"] == age_bin][attr], p) for age_bin in age_bin_labels]

        if (p == 50):   ## midpoint
            rsqr = Get_RSquared(x, y)

        fit = np.polyfit(x, y, polynomial_order)
        poly = np.poly1d(fit)
        age_bin_percentiles.append(poly)

#        if (polynomial_order == 1) and (p == 50): # if this is
#         if (p == 50):  # if this is the median value, remember the equation
#             equation = poly

    return age_bins, age_bin_labels, age_bin_percentiles, age_bin_percentiles_df, percentiles, rsqr

def DefineColours():
    # Set the colormap for the percentile range
    colors = [  # ver 4
        [0.8, 0.0, 0.0, 0.65],  # Red
        [1.0, 0.5, 0.0, 0.6],  # Orange
        [1.0, 1.0, 0.0, 0.5],  # Yellow
        [0.0, 1.0, 0.0, 0.4],  # Green
        [0.0, 0.7, 1.0, 0.5]  # Blue
    ]

    for i in range(1, 6):
        x = colors[5 - i]
        colors = np.vstack([colors, x])

    return colors

def FilterDataBySex(df, sex):
    if (sex != ''):
        # we are filtering by sex.  include patient_sex parameter
        filtered_data = df.query(f'patient_age >= {min_age} and patient_age <= {max_age} and patient_sex == "{sex}"' + " " + extra_filter, inplace = False)
        filtered_data = filtered_data[['patient_age', 'patient_sex', attr]]
    else:
        # we are not filtering by sex.  do not include patient_sex parameter
        filtered_data = df.query(f'patient_age >= {min_age} and patient_age <= {max_age} ' + " " + extra_filter, inplace=False)
        filtered_data = filtered_data[['patient_age', attr]]

    return filtered_data

def DefineCurrentPatient(sex, mid_point):
    # Create the current patient's data (this is optional if you want to point to a particular patient)
    current_patient = {
        "patient_sex": sex,
        "patient_age": 35,
        attr: mid_point, # this value is an arbitrary y-axis value (eg. liver_volume of 2000)
    }

    return current_patient

def SetLowHighFilterValues(bAlwaysUseRecommended):
    extra_filter = ""
    print("Column " + attr + " has the following attributes.")
    stat_min = df[attr].min()
    stat_max = df[attr].max()
    stat_mean = df[attr].mean()
    stat_desired_percentile = round(df[attr].quantile(0.98))
    print("Min: ", str(round(stat_min)))
    print("Max: ", str(round(stat_max)))
    print("Mean: ", str(round(stat_mean)))
    print("Number of zeros: ", (df[attr] == 0).sum())
    print("Number of records (rows): ", df[attr].count())
    print("98th Percentile: ", str(stat_desired_percentile))

    # round the filter to the nearest 10 or 100
    if (stat_max < 10):
        guess_low_filter = stat_min
        guess_high_filter = stat_max
    elif (stat_max < 1000):
        guess_low_filter = round((stat_mean - (stat_desired_percentile / 2)) / 10.0) * 10
        guess_high_filter = round((stat_mean + (stat_desired_percentile / 2)) / 10.0) * 10
    elif (stat_max >= 1000):
        guess_low_filter = round((stat_mean - (stat_desired_percentile / 2)) / 100.0) * 100
        guess_high_filter = round((stat_mean + (stat_desired_percentile / 2)) / 100.0) * 100

    if (bAlwaysUseRecommended == False):
        user_filter_string = input("Enter low, high filter values to use? eg. " + str(guess_low_filter) + ", " + str(
            guess_high_filter) + " (blank is no filtering, (r) uses these recommended values once, (a) uses these recommended values always)? ")

        if (user_filter_string == 'a'): # we want to use filters for ALL subsequent processings
            bAlwaysUseRecommended = True

    if (bAlwaysUseRecommended == True):
        user_filter_string = 'r'

    if (user_filter_string == 'r'):  # recommended filters (using guess)
        low_filter = guess_low_filter
        high_filter = guess_high_filter
        mid_point = ((high_filter - low_filter) / 2) + low_filter
        extra_filter = "and " + attr + " >= " + str(low_filter) + " and " + attr + " <= " + str(high_filter)
    elif (user_filter_string != ''):  # user defined filters
        user_filter_split = user_filter_string.split(", ")
        low_filter = float(user_filter_split[0])
        high_filter = float(user_filter_split[1])
        mid_point = ((high_filter - low_filter) / 2.0) + low_filter
        extra_filter = "and " + attr + " >= " + str(low_filter) + " and " + attr + " <= " + str(high_filter)
    else:
        mid_point = df[attr].mean()
        extra_filter = ""

    return bAlwaysUseRecommended, extra_filter, mid_point;


def SelectSex():
    uin2 = ''
    chart_both_sexes = False
    while (uin2 != 'f') and (uin2 != 'm') and (uin2 != 'c') and (uin2 != 'b'):
        uin2 = input("Enter sex: m for male, f for female, c for combined, b for both on separate charts: ")
        if (uin2 == 'm'):
            sex = 'male'
        elif uin2 == 'f':
            sex = 'female'
        elif uin2 == 'b':
            sex = 'male'
            chart_both_sexes = True
        elif uin2 == 'c':
            sex = ''

    return sex, chart_both_sexes

def SelectColumnFromFile(df, column_index):
    if (column_index <= len(df.columns)):
        attr = df.columns[column_index - 1]
        print(f'Using column {attr}')
    else:
        attr = ""
        print(f'### error: Workbook only has {len(df.columns)} columns.  Column {column_index} is out of bounds')

    return attr

def GetColumns():
    # allow user to filter on the keyword to find in the worksheet
    # eg. if a column is titled "lung_volume", entering the keyword "volume" will find lung_volume, chest_volume, etc
    keyword_search = input("Enter keyword to search for in column titles (eg. volume): ")
    for i in range(len(df.columns)):
        if keyword_search in df.columns[i]:
            if (i % 5 == 0):  # print 5 columns per row (across the screen)
                print(f'{i + 1}. {df.columns[i]}, ')
            else:
                print(f'{i + 1}. {df.columns[i]}, ', end=" ")

    print();
    uin1 = input("Which column numbers to use for y-axis of the graph, eg. 10, 22, 35? ")
    my_columns = [int(n) for n in uin1.split(", ")] #uin1.split(", ")

    return my_columns


# if reading an excel file, select the specific sheet
def SelectExcelSheet(sheet_tabs):
    print("when opening an Excel file, we need to select the specific Worksheet.  Please select the desired Worksheet")
    for i in range(len(sheet_tabs)):
        print(f'{i+1}. {sheet_tabs[i]}')

    sheet_index = 0
    while (sheet_index == 0):
        sheet_index = input('Enter Worksheet number to analyse: ')

    return sheet_tabs[int(sheet_index)-1]

# Build a list of files from the current directory (csv, xlsx)
def BuildFileList():
    # Define paths
    curr_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.dirname(curr_dir)
    input_dir = curr_dir  # os.path.join(parent_dir, "Data", "Historical_01", "Input")
    output_dir = curr_dir  # os.path.join(parent_dir, "Data", "Historical_01", "Output")
    my_sheet_tab = ""       # assume we don't have a worksheet (which only might apply to an excel file)

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # find all csv files in the current directory to show to the user
    filelistcsv = glob.glob('**/*.csv', recursive=True)
    filelistxlsx = glob.glob('**/*.xlsx', recursive=True)
    if (len(filelistcsv) > 0) or (len(filelistxlsx) > 0):
        print ("Found the following files for processing:")
        # build list of csv files
        for i in range(len(filelistcsv)):
            if (i % 5 == 0):
                print(f'{i+1}. {filelistcsv[i]}, ')
            else:
                print(f'{i+1}. {filelistcsv[i]}, ', end=" ")

        print();
        print();

        # build list of csv files
        for j in range(len(filelistxlsx)):
            if (j % 5 == 0):
                print(f'{chr(97+j)}. {filelistxlsx[j]}, ')
            else:
                print(f'{chr(97+j)}. {filelistxlsx[j]}, ', end=" ")

        print();
        # ft = DataFileType.NONE
        i = input("which file do you want to use (default 1)? ")
        if (i == ''):
            i = "1"

        if (i >= '1') and (i <= '9'):
            # csv filenames are indexed with '1', '2', etc
            filename = filelistcsv[int(i)-1]

            # Import spreadsheet data from a csv file.  First column must be column_name
            df = pd.read_csv(os.path.join(input_dir, filename))
            # ft = DataFileType.CSV
        else:
            # xlsx filenames are indexed with 'a', 'b', etc
            filename = filelistxlsx[ord(i)-97]

            # Import spreadsheet data from a csv file.  First column must be column_name
            sheet_tabs = load_workbook(os.path.join(input_dir, filename), read_only = True, keep_links = False).sheetnames
            my_sheet_tab = SelectExcelSheet(sheet_tabs)
            df = pd.read_excel(os.path.join(input_dir, filename), sheet_name=my_sheet_tab)
    else:
        print("No .csv files in the following directory.  Closing program" + curr_dir)
        exit(0)

    return output_dir, my_sheet_tab, df

############################## MAIN #########################################

# clear the terminal window
os.system('cls');

# Filter the data by patient_sex, patient_age
# note: uses current_patient (above) to filter by sex
min_age = 15
max_age = 100
bAlwaysUseRecommended = False   # don't always use recommended filter values
uin = input("Enter maximum polynomial order to calculate (1 = linear, 2 = 2nd order polynomial, etc) : ")
max_polynomial_order = int(uin)

colors = DefineColours()

output_dir, my_sheet_tab, df = BuildFileList()
my_columns = GetColumns()
sex, chart_both_sexes = SelectSex()

for i in my_columns:
    attr = SelectColumnFromFile(df, i)

    if (attr != ""):
        if chart_both_sexes:
            sexes_count = 2
            sex = "male"
        else:
            sexes_count = 1

        while (sexes_count > 0):
            polynomial_order = 1
            bAlwaysUseRecommended, extra_filter, mid_point = SetLowHighFilterValues(bAlwaysUseRecommended)
            current_patient = DefineCurrentPatient(sex, mid_point)
            filtered_data = FilterDataBySex(df, sex)

            if (len(filtered_data) > 0):    # we have filtered data
                while (polynomial_order <= max_polynomial_order): # for each polynomial (1 = linear, 2 = 2nd order poly, etc)
                    age_bins, age_bin_labels, age_bin_percentiles, age_bin_percentiles_df, percentiles, rsqr = CalculatePercentiles(filtered_data)

                    x_plot = np.linspace(0, max_age, 500)
                    # remove _ from column name and word capitalise (eg. liver_volume = Liver Volume)
                    attr_formatted = attr.replace('_', ' ').title()

                    plt = PlotPercentileRanges(polynomial_order)
                    plt = PlotScatterPlot(plt, filtered_data)
                    rsqr2, plt = PlotTrendLine(plt) ### debug - don't need rsqr2 (we have rsqr from above)
                    #plt = PlotFakePatient(plt, current_patient) # we made up a patient (above)
                    PlotEquation(plt, rsqr)
                    plt = PlotTrimmings(my_sheet_tab)

                    if (my_sheet_tab != ""):
                        plt.savefig(f"{output_dir}/{my_sheet_tab} {attr} poly {polynomial_order}_{sex}.png")
                    else:
                        plt.savefig(f"{output_dir}/{attr} poly {polynomial_order}_{sex}.png")

                    plt.show()
                    plt.close() # clear the plot read to generate a second plot

                    polynomial_order += 1
            else:
                print(f'no data for {sex}')

            # if we started with male and we are charting both sexes, end with female
            if (sexes_count == 2):
                sex = "female"

            sexes_count -= 1
