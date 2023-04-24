import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import statsmodels.api as sm
import statsmodels.formula.api as smf
import os
import glob # for directory parsing
from openpyxl import load_workbook  # attempt 2 at fast loading excel worksheets
from datetime import date


##########################################################################################
# Instructions Summary for the AMRI Chart Generator Script:
#
# 1. Ensure that the required input files (CSV or Excel) are located in the 'Input' directory.
# 2. The script will display a list of available CSV and Excel files for processing.
# 3. Select the desired file by entering the corresponding number or letter (CSV files are numbered, Excel files are lettered).
# 4. Select the columns you want to use for the X and Y axes. Columns are numbered for user convenience.
# 5. Choose the sex(es) to be charted in the analysis by entering 'm' for male, 'f' for female, 'b' for both sexes, 'c' for combined sexes, or 'a' for all charts.
# 6. Enter a filter range for the data by specifying 'a' for always, 'n' for never, followed by the minimum and maximum values (e.g., a15,100).
# 7. Confirm if you would like to display the charts by entering 'Y' for yes or 'N' for no.
# 8. The script will process the data, compute quantile regression, and generate the charts based on your input.
# 9. The generated charts will be saved in the 'Output' directory.
##########################################################################################



# Define constants
BOTH_SEXES = " "
STR_GENERIC_FIELD = "strFieldName"

# Function to compute the Quantile Regression for the given dataset (X, y) and the specified quantiles.
def compute_QuantReg(bX, y, quantiles, model_type='linear'):

    df = pd.concat([X, y], axis=1)
    df.columns = ['X', 'y']
    results = []
    for q in quantiles:
        if model_type == 'linear':
            model = smf.quantreg('y ~ X', df).fit(q=q)
        elif model_type == 'polynomial':
            degree = 2
            model = smf.quantreg(f'y ~ X + I(X**{int(degree)})', df).fit(q=q)
        else:
            raise ValueError("Invalid model type. Choose either 'linear' or 'polynomial'.")
        results.append(model)

    return results


# Function to exclude outlier data points from the input X and y data by applying the IQR method.
def exclude_outliers(X, y):
    m = 2.8 # adjust this to change outlier tolerance
    Q1_y = y.quantile(0.25)
    Q3_y = y.quantile(0.75)
    IQR_y = Q3_y - Q1_y
    if X.dtype != 'int64':
        Q1_x = X.quantile(0.25)
        Q3_x = X.quantile(0.75)
        IQR_x = Q3_x - Q1_x
    else:
        IQR_x = None
    outliers_y = ((y < (Q1_y - m * IQR_y)) | (y > (Q3_y + m * IQR_y)))
    if IQR_x is not None:
        outliers_x = ((X < (Q1_x - m * IQR_x)) | (X > (Q3_x + m * IQR_x)))
    else:
        outliers_x = pd.Series([False] * len(X))
    outliers = outliers_y | outliers_x
    X_no_outliers = X.loc[~outliers]
    y_no_outliers = y.loc[~outliers]
    
    return X_no_outliers, y_no_outliers, outliers


# Creates a scatter plot of the Quantile Regression results for the given dataset (X, y), sex, filter range, and quantiles.
def QuantReg_scatterplot(bShowCharts, sex, strFilter, X, y, results, quantiles, colors, print_summary=False):
    X_no_outliers, y_no_outliers, outliers = exclude_outliers(X, y)
    x_label = X.name.replace('_', ' ').title()
    y_label = y.name.replace('_', ' ').title()

    fig, ax = plt.subplots(figsize=(10, 6))
    x_values = np.linspace(min(X), max(X), 100)
    ax.scatter(X_no_outliers, y_no_outliers, alpha=0.5, s=22, label="Data")
    ax.scatter(X[outliers], y[outliers], alpha=0.5, s=22, label="Outliers", c='r', marker='x')
    # Fit and plot OLS linear regression as well
    cleaned_data = pd.concat([X_no_outliers, y_no_outliers], axis=1).dropna()
    X_cleaned = cleaned_data[X_no_outliers.name]
    y_cleaned = cleaned_data[y_no_outliers.name]
    ols_model = sm.OLS(y_cleaned, sm.add_constant(X_cleaned)).fit()
    ax.plot(x_values, ols_model.params[0] + ols_model.params[1] * x_values, label='OLS Linear Regression', c='k', linestyle='dashed', linewidth=1.5)
    for i, (q, model) in enumerate(zip(quantiles, results)):
        color = colors[i % len(colors)]
        if len(model.params) == 2:
            y_values = model.params[0] + model.params[1] * x_values
            equation = f"{model.params[1]:.2f}x + {model.params[0]:.2f}"
        elif len(model.params) == 3:
            y_values = model.params[0] + model.params[1] * x_values + model.params[2] * (x_values ** 2)
            equation = f"{model.params[2]:.2f}x^2 + {model.params[1]:.2f}x + {model.params[0]:.2f}"
        ax.plot(x_values, y_values, color=color, label=f"{int(q*100)}th Pctl, {equation}")
        if print_summary:
            print(f"{int(q * 100)}th Quantile Regression Model Summary")
            print("=" * 40)
            print(model.summary())
    ax.set_xlabel('X: '+x_label, fontsize=10)
    ax.set_ylabel('Y: '+y_label, fontsize=10)
    ax.legend(fontsize=8)
    plt.text(1.0, 1.0, f"OLS R^2: {ols_model.rsquared:.4f}", fontsize=10, transform=ax.transAxes,horizontalalignment='right', verticalalignment ='bottom')
    plt.grid(which='major', linestyle='-', color='black')
    plt.grid(which='minor', linestyle='--')
    plt.minorticks_on()
    plt.title(f"{sex.title()} {x_label} vs {y_label} {strFilter} @{date.today()}")
    plt.tight_layout()
    if (sex != BOTH_SEXES):
        plt.savefig(f'{x_label} vs {y_label} ({sex}).png')
    else:
        plt.savefig(f'{x_label} vs {y_label}.png')

    if (bShowCharts == True):
        plt.show()
    plt.close()


#########################################################################################################

# Function that allows the user to select a column from the given DataFrame (df) based on the provided column_index.
def SelectNextColumnFromFile(df, column_index):
    if (column_index <= len(df.columns)):
        attr = df.columns[column_index - 1]
        print(f'Using column {attr}')
    else:
        attr = ""
        print(f'### error: Workbook only has {len(df.columns)} columns.  Column {column_index} is out of bounds')

    return attr


# Function to prompt the user to enter the columns to be processed in pairs (X, Y).
def GetColumns():
    # allow user to filter on the keyword to find in the worksheet
    # eg. if a column is titled "lung_volume", entering the keyword "volume" will find lung_volume, chest_volume, etc
    bCommaDetected = False

    while (bCommaDetected == False):
        print()
        print ("="*40)
        uin1 = input("Please choose the columns you would like to use for x-axis and y-axis of the graph?\n"
                     "Eg. 10, 22, 35, 41    Would chart 10 vs 22, then 35 vs 41\n"
                     "ORr ?XXXX for list of columns,\n"
                     "Eg. ?vol    Would find fields containing the word 'vol'\n"
                     "Your input:")
        if ("?" in uin1):
            keyword_search = uin1.replace("?", "")

            for i in range(len(df.columns)):
                if keyword_search in df.columns[i]:
                    if (i % 5 == 0):  # print 5 columns per row (across the screen)
                        print(f'{i + 1}. {df.columns[i]}, ')
                    else:
                        print(f'{i + 1}. {df.columns[i]}, ', end=" ")

            print()
        elif ("," in uin1):
            my_columns = [int(n) for n in uin1.split(",")] #uin1.split(", ")
            bCommaDetected = True
        else:
            print("Error: you must have at least two column indexes, separated by a comma, eg. 3,4")


    return my_columns


# if reading an excel file, select the specific sheet
def SelectExcelSheet(sheet_tabs):
    print()
    print ("="*40)
    print("For an excel file, we need to select the specific Worksheet:")
    for i in range(len(sheet_tabs)):
        print(f'{i+1}. {sheet_tabs[i]}')

    sheet_index = 0
    while (sheet_index == 0):
        sheet_index = input('Your input: ')

    return sheet_tabs[int(sheet_index)-1]


# Function to build a list of file paths for the data files to be used in the analysis (divided into csv and xlsx).
def BuildFileList():
    # Define paths
    curr_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.dirname(curr_dir)
    input_dir = os.path.join(parent_dir, "Data", "Historical_01", "Input")
    output_dir = os.path.join(parent_dir, "Data", "Historical_01", "Output")
    # input_dir = curr_dir
    # output_dir = curr_dir

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # find all csv files in the current directory to show to the user
    filelistcsv = glob.glob('**/*.csv', recursive=True)
    filelistcsv = [os.path.basename(i) for i in filelistcsv]
    filelistxlsx = glob.glob('**/*.xlsx', recursive=True)
    filelistxlsx = [os.path.basename(i) for i in filelistxlsx]

    if (len(filelistcsv) > 0) or (len(filelistxlsx) > 0):
        print()
        print ("="*40)
        print ("Found the following files for processing:")
        # build list of csv files
        print("CSV Files:")
        for i in range(len(filelistcsv)):
            if (i % 5 == 0):
                print(f'{i+1}. {filelistcsv[i]}, ')
            else:
                print(f'{i+1}. {filelistcsv[i]}, ', end=" ")

        # build list of csv files
        print("\nExcel Files:")
        for j in range(len(filelistxlsx)):
            if (j % 5 == 0):
                print(f'{chr(97+j)}. {filelistxlsx[j]}, ')
            else:
                print(f'{chr(97+j)}. {filelistxlsx[j]}, ', end=" ")

        print()
        # ft = DataFileType.NONE
        i = input("Which file do you want to use (Default = 1)?\n"
                  "Your input: ")
        if (i == ''):
            i = "1"

        if (i >= '1') and (i <= '9'):
            # csv filenames are indexed with '1', '2', etc
            filename = filelistcsv[int(i)-1]

            # Import spreadsheet data from a csv file.  First column must be column_name
            df = pd.read_csv(os.path.join(input_dir, filename))
            # ft = DataFileType.CSV
        else:
            # xlsx filenames are indexed with 'a', 'b', etc
            filename = filelistxlsx[ord(i)-97]

            # Import spreadsheet data from a csv file.  First column must be column_name
            sheet_tabs = load_workbook(os.path.join(input_dir, filename), read_only = True, keep_links = False).sheetnames
            my_sheet_tab = SelectExcelSheet(sheet_tabs)
            df = pd.read_excel(os.path.join(input_dir, filename), sheet_name=my_sheet_tab)
    else:
        print("No .csv files in the following directory.  Closing program" + curr_dir)
        exit(0)

    return df

# Function to prompt the user to choose the sexes to be charted in the analysis.
def GetSexesToChart():
    chart_sexes = []
    while (len(chart_sexes) == 0):
        print()
        print ("="*40)
        uin = input(
            "Please choose the sexes to be charted in the analysis:\n"
            "(m)ale\n"
            "(f)emale\n"
            "(b)oth sexes (male and female separately)\n"
            "(c)ombined sexes (all data combined)\n"
            "(a)ll charts (both, combined, male, and female)\n"
            "Your input: "
            ).lower()
        if (uin == 'm'):
            chart_sexes.append('male')
        elif (uin == 'f'):
            chart_sexes.append('female')
        elif (uin == 'b'):
            chart_sexes.append('male')
            chart_sexes.append('female')
        elif (uin == 'c'):
            chart_sexes.append(BOTH_SEXES)
        elif (uin == 'a'):
            chart_sexes.append(BOTH_SEXES)
            chart_sexes.append('male')
            chart_sexes.append('female')

    return chart_sexes


# Function to return a query string for a generic field (STR_GENERIC_FIELD)
# this can then be substitited for the correct column as we process each column
# eg. STR_GENERIC_FIELD -> Liver HU mean,
#     STR_GENERIC_FIELD -> Kidney HU mean
# note: this conversion is done outside of this function
def GetFilterRange(bFilterAlways, bFilterNever, strFieldName):
    strFilter = ""
    print()
    print ("="*40)
    uin = input(
            f"Enter x-axis filter range for {strFieldName} using the following format:\n"
            "(amin, max) or (nmin, max)\n"
            "a = always apply filter\n"
            "n = never apply filter\n"
            "Example: a1.5, 2\n"
            "Your input: "
            ).lower()
    if (uin != ""):
        if (uin[0] == 'a'):
            bFilterAlways = True
            uin = uin.replace('a', '')
        elif (uin[0] == 'n'):
            bFilterNever = True
            uin = uin.replace('n', '')

        temp = []
        if ("," in uin): # we have min, max
            temp = uin.split(",")
            fFilterMin = float(temp[0])
            fFilterMax = float(temp[1])
        else:
            temp.append(uin)

        if (len(temp) > 0) and (bFilterNever == False): # we have a low or high parameter
            strFilter = f"{STR_GENERIC_FIELD} >= {temp[0]}"
            if (len(temp) > 1):
                strFilter = strFilter + f" and {STR_GENERIC_FIELD} <= {temp[1]}"

            print(strFilter)  ### debug

    return bFilterAlways, bFilterNever, strFilter


#########################################################################################################
#########################################################################################################
#########################################################################################################

# set static data
# Example usage
quantiles = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]
colors = [  # ver 5
    [0.8, 0.0, 0.0, 0.65], # Red
    [1.0, 0.5, 0.0, 0.6],  # Orange
    [1.0, 1.0, 0.0, 0.5],  # Yellow
    [0.0, 1.0, 0.0, 0.4],  # Green
    [0.0, 0.7, 1.0, 0.5],  # Blue
    [0.0, 1.0, 0.0, 0.4],  # Green
    [1.0, 1.0, 0.0, 0.5],  # Yellow
    [1.0, 0.5, 0.0, 0.6],  # Orange
    [0.8, 0.0, 0.0, 0.65]  # Red
]

# initialise variables used in the program
bAlwaysUseFilter = False
bNeverUseFilter = False
bFilterMin = 0.0
bFilterMax = 0.0
strGenericFilter = ""

# clear the terminal window
os.system('cls')

# Load file
#df = pd.read_csv('/home/richard/Richard/RR-personal-repo/Data/Historical_01/Input/RR Image Similarity Parameters with Patient Data - Test.csv') # User to change

df = BuildFileList()
my_columns = GetColumns()

# Prompt the user to enter the columns to be processed in pairs (X, Y).
chart_sexes = GetSexesToChart()
print()
print ("="*40)
uin = input("Show charts (Y/N) : ")
if (uin == "Y") or (uin == "y"):
    bShowCharts = True
else:
    bShowCharts = False

for i in range(len(my_columns)):
    if (i%2 == 1):  # skip every second column (ie. we process columns in pairs, so we process 0,1 2,3, 4,5)
        continue

    # we converted columns to integers for user input convenience.
    # now we need to convert them back again to support the code
    col_X_name = df.columns[my_columns[i]-1]
    col_y_name = df.columns[my_columns[i+1]-1]

    # determine if we need to filter the data
    # we can always use this filter or never use this filter
    if (bAlwaysUseFilter == False) and (bNeverUseFilter == False):
        bAlwaysUseFilter, bNeverUseFilter, strGenericFilter = GetFilterRange(bAlwaysUseFilter, bNeverUseFilter, col_X_name)

    for sex in chart_sexes:
        if (sex != BOTH_SEXES): # include all records
            # grouped = df.groupby('patient_sex')
            # df1 = grouped.get_group(sex)
            full_query = f'patient_sex == "{sex}"'
            if strGenericFilter != '':
                full_query = full_query + " and " + strGenericFilter.replace(STR_GENERIC_FIELD, col_X_name)
            df1 = df.query(full_query)
        elif (strGenericFilter != ''):
            df1 = df.query(strGenericFilter.replace(STR_GENERIC_FIELD, col_X_name))
        else:
            df1 = df

        X = df1[col_X_name]
        y = df1[col_y_name]


        # Get quantile regression results
        results = compute_QuantReg(X,y, quantiles, model_type='linear')

        # Plot Quantile regression results, also conducts outlier exclusion internally
        # note: sex is the sex of the patient (visual only for title and filename)
        # note: strGenericFilter shows the filter range.  'X' shortens the filter to X >= 45 and X <= 55
        QuantReg_scatterplot(bShowCharts, sex, strGenericFilter.replace(STR_GENERIC_FIELD, 'X'), X, y, results, quantiles, colors=colors)
